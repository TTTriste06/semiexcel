import pandas as pd
from typing import Tuple
from openpyxl.styles import PatternFill

def standardize(val):
    if val is None:
        return ""
    return str(val).strip().replace("\u3000", " ").strip('\'"“”‘’')

def append_forecast_unmatched_to_summary_by_keys(summary_df: pd.DataFrame, forecast_df: pd.DataFrame, ws=None) -> Tuple[pd.DataFrame, list]:
    """
    将预测中未匹配的记录补充进汇总表，并可选在 Excel 中标红。
    
    参数:
    - summary_df: 汇总 DataFrame（含“晶圆品名”、“规格”、“品名”列）
    - forecast_df: 原始预测 DataFrame（含“产品型号”、“生产料号”列）
    - ws: 可选 openpyxl 工作表对象（用于填充颜色）

    返回:
    - summary_df: 更新后的 DataFrame
    - added_indices: 新增行在 Excel 中的位置（用于后续标记）
    """
    forecast_cols = [col for col in forecast_df.columns if "预测" in col]
    needed_cols = ["产品型号", "生产料号"] + forecast_cols
    forecast_subset = forecast_df[needed_cols].copy()

    unmatched = forecast_subset[~forecast_subset["生产料号"].isin(summary_df["品名"])].copy()

    unmatched["规格"] = unmatched["产品型号"]
    unmatched["品名"] = unmatched["生产料号"]
    unmatched["晶圆品名"] = ""

    summary_cols = summary_df.columns
    new_rows = unmatched[["晶圆品名", "规格", "品名"] + forecast_cols if forecast_cols else []]

    for col in summary_cols:
        if col not in new_rows.columns:
            new_rows[col] = ""

    new_rows = new_rows[summary_cols]

    # 新增行 Excel 位置（用于 openpyxl 填色）
    start_idx = len(summary_df)
    added_indices = list(range(start_idx + 2, start_idx + 2 + len(new_rows)))  # header 在第2行

    summary_df = pd.concat([summary_df, new_rows], ignore_index=True)

    # ✅ 标红 Excel 对应的新增行
    if ws is not None:
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        for row_idx in added_indices:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill

    return summary_df, added_indices
