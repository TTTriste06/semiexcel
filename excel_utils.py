import pandas as pd
import streamlit as st
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.cell import MergedCell


def standardize(val):
    """
    将输入标准化为可比较的字符串：
    - 转字符串
    - 去除首尾空格
    - 去除单引号/双引号（英文和中文）
    - 统一半角/全角空格
    """
    if val is None:
        return ''
    
    val = str(val).strip()
    
    # 去掉包裹的引号（包括中英文单引号和双引号）
    val = val.strip('\'"“”‘’')  # 含中文引号
    
    # 替换全角空格为半角空格
    val = val.replace('\u3000', ' ')

    return val







def clean_df(df):
    df = df.fillna("")  # 处理真正的 NaN/None
    df = df.applymap(lambda x: "" if str(x).strip().lower() == "nan" else str(x).strip() if isinstance(x, str) else x)
    return df


def clear_nan_cells(ws):
    """
    遍历指定工作表，将 'nan'、'NaN'、None 和空字符串统一清空。
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip().lower() == "nan":
                cell.value = ""


def clean_key_fields(df, field_map):
    for col in [field_map["规格"], field_map["品名"], field_map["晶圆品名"]]:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace(r"\s+", "", regex=True)  # 移除所有空白字符（含空格、全角空格、Tab、换行）
            .str.replace(r"[\u200b\u200e\u200f]", "", regex=True)  # 清除不可见字符
            .str.strip()
        )
    return df










def adjust_column_width(writer, sheet_name, df):
    """
    自动调整 Excel 工作表中各列的宽度以适应内容长度。

    参数:
    - writer: pandas 的 ExcelWriter 对象
    - sheet_name: 要调整的工作表名称
    - df: 对应写入工作表的 DataFrame 数据
    """
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns, 1):
        # 获取该列中所有字符串长度的最大值
        max_content_len = df[col].astype(str).str.len().max()
        header_len = len(str(col))
        column_width = max(max_content_len, header_len) * 1.2 + 8
        worksheet.column_dimensions[get_column_letter(idx)].width = min(column_width, 50)

def adjust_column_width_ws(ws):
    """
    根据每个单元格内容长度，自动调整 openpyxl Worksheet 的列宽。
    """
    column_widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if cell is not None:
                width = len(str(cell))
                if i in column_widths:
                    column_widths[i] = max(column_widths[i], width)
                else:
                    column_widths[i] = width

    for i, width in column_widths.items():
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = width + 88  # 可调节 +2 缓冲








def merge_header_for_summary(ws, df, label_ranges):
    """
    给指定汇总列添加顶部合并行标题（如“安全库存”“未交订单”）

    参数:
    - ws: openpyxl worksheet
    - df: summary DataFrame
    - label_ranges: dict，键是标题文字，值是列名范围元组，如：
        {
            "安全库存": (" InvWaf", " InvPart"),
            "未交订单": ("总未交订单", "未交订单数量_2025-08")
        }
    """

    # 插入一行作为新表头（原表头往下挪）
    ws.insert_rows(1)
    header_row = list(df.columns)

    for label, (start_col_name, end_col_name) in label_ranges.items():
        if start_col_name not in header_row or end_col_name not in header_row:
            continue

        start_idx = header_row.index(start_col_name) + 1  # Excel index starts from 1
        end_idx = header_row.index(end_col_name) + 1

        col_letter_start = get_column_letter(start_idx)
        col_letter_end = get_column_letter(end_idx)

        merge_range = f"{col_letter_start}1:{col_letter_end}1"
        ws.merge_cells(merge_range)
        cell = ws[f"{col_letter_start}1"]
        cell.value = label
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

def get_column_index_by_name(ws, target_name: str, header_row: int = 1):
    """
    从 Excel sheet 的指定表头行中查找列名对应的列号（1-based）
    """
    for cell in ws[header_row]:
        if str(cell.value).strip() == target_name:
            return cell.column
    return None


def merge_duplicate_product_names(summary_df: pd.DataFrame) -> pd.DataFrame:
    """
    合并 '汇总' 表中重复的品名（按品名分组），选用第一行的 晶圆品名 和 规格，合并其数值列。
    """
    # 确保必要列存在
    required_cols = ["晶圆品名", "规格", "品名"]
    for col in required_cols:
        if col not in summary_df.columns:
            raise ValueError(f"缺少必要列：{col}")

    # 识别数值列（排除主键列）
    value_cols = [col for col in summary_df.columns if col not in required_cols]

    # 分组合并数值列
    grouped = summary_df.groupby("品名", sort=False)

    merged_rows = []

    for name, group in grouped:
        if len(group) == 1:
            merged_rows.append(group.iloc[0])
        else:
            # 取第一行的 晶圆品名 和 规格
            base_row = group.iloc[0][required_cols].copy()
            summed_values = group[value_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum()
            merged_row = pd.concat([base_row, summed_values])
            merged_rows.append(merged_row)

    # 合并所有结果
    merged_df = pd.DataFrame(merged_rows)

    # 保证列顺序与原始一致
    return merged_df[summary_df.columns]

def merge_duplicate_rows_by_key(df: pd.DataFrame, field_map: dict, verbose=True) -> pd.DataFrame:
    """
    合并给定表格中 '规格' + '品名' + '晶圆品名' 相同的行：
    - 数值列求和
    - 其他字段取第一行
    - 主键字段来自 field_map

    增加 verbose 输出用于调试未合并成功的情况
    """

    key_cols = [field_map["规格"], field_map["品名"], field_map["晶圆品名"]]

    for col in key_cols:
        if col not in df.columns:
            raise ValueError(f"缺少必要列：{col}")

    # 主键列清洗
    for col in key_cols:
        df[col] = (
            df[col]
            .astype(str)
            .str.strip()
            .str.replace(r"\s+", "", regex=True)
            .str.replace(r"[\u200b\u200e\u200f\n\r]", "", regex=True)
        )

    # 调试输出重复主键组合
    dup_keys = df.groupby(key_cols).size().reset_index(name="count")
    dup_keys = dup_keys[dup_keys["count"] > 1]

    if verbose and not dup_keys.empty:
        st.warning(f"⚠️ 检测到 {len(dup_keys)} 个重复主键组合，准备合并：")
        for idx, row in dup_keys.iterrows():
            key_values = tuple(row[col] for col in key_cols)
            st.write(f"🔁 主键组：{key_values}")
            st.dataframe(df[
                (df[key_cols[0]] == key_values[0]) &
                (df[key_cols[1]] == key_values[1]) &
                (df[key_cols[2]] == key_values[2])
            ])

    # 数值列合并
    value_cols = [col for col in df.columns if col not in key_cols and pd.api.types.is_numeric_dtype(df[col])]
    grouped = df.groupby(key_cols, sort=False)
    merged_rows = []

    for keys, group in grouped:
        if len(group) == 1:
            merged_rows.append(group.iloc[0])
        else:
            base_row = group.iloc[0][df.columns.difference(value_cols)].copy()
            summed_values = group[value_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum()
            merged_row = pd.concat([base_row, summed_values])
            merged_rows.append(merged_row)

    merged_df = pd.DataFrame(merged_rows)

    # 保持列顺序一致
    return merged_df[df.columns]


# 💡 自动按模块排序汇总列：安全库存 → 未交订单 → 预测 → 其他
def reorder_summary_columns(df):
    col_list = df.columns.tolist()

    # 区分模块
    safe_cols = [col for col in col_list if "Inv" in col]
    unfulfilled_cols = [col for col in col_list if "未交订单" in col or col == "总未交订单"]
    forecast_cols = [col for col in col_list if "预测" in col]
    base_cols = ["晶圆品名", "规格", "品名"]
    other_cols = [col for col in col_list if col not in base_cols + safe_cols + unfulfilled_cols + forecast_cols]

    # 组合：基础列 + 安全库存 + 未交订单 + 预测 + 其他
    new_order = base_cols + safe_cols + unfulfilled_cols + forecast_cols + other_cols
    return df[new_order]











def mark_unmatched_keys_on_sheet(ws, unmatched_keys, wafer_col=1, spec_col=2, name_col=3):
    """
    在 openpyxl 工作表中标红未匹配的行（通过主键匹配），对空值/None做标准化处理。

    参数:
    - ws: openpyxl worksheet 对象
    - unmatched_keys: list of (晶圆品名, 规格, 品名) 元组
    - wafer_col, spec_col, name_col: 表示主键列在 sheet 中的列号（从1开始）
    """
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    unmatched_set = set(
        tuple(standardize(x) for x in key)
        for key in unmatched_keys
    )

    for row in range(2, ws.max_row + 1):  # 从第2行开始
        wafer = standardize(ws.cell(row=row, column=wafer_col).value)
        spec = standardize(ws.cell(row=row, column=spec_col).value)
        name = standardize(ws.cell(row=row, column=name_col).value)

        if (wafer, spec, name) in unmatched_set:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill

def mark_unmatched_keys_on_name(ws, unmatched_keys, name_col=3):
    """
    在 openpyxl 工作表中标红未匹配的品名行。

    参数:
    - ws: openpyxl worksheet 对象
    - unmatched_keys: list of 品名
    - name_col: 品名所在列（从 1 开始）
    """
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # 统一标准化格式
    unmatched_set = set(standardize(key) for key in unmatched_keys)

    for row in range(2, ws.max_row + 1):  # 从第2行开始处理
        name = standardize(ws.cell(row=row, column=name_col).value)

        if name in unmatched_set:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill


def mark_keys_on_sheet(ws, key_set, key_cols=(1, 2, 3)):
    """
    在工作表中标黄匹配 key_set 中的行，基于主键列匹配。

    参数:
    - ws: openpyxl worksheet
    - key_set: set of tuple，例如 {("晶圆品名", "规格", "品名"), ...}
    - key_cols: 表示主键所在的列号 (从1开始)，默认是 (1, 2, 3) 对应“晶圆品名”, “规格”, “品名”
    """
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    def standardize(val):
        if val is None:
            return ''
        val = str(val)
        val = val.replace('\u3000', ' ')  # 全角空格
        val = re.sub(r"[\"'‘’“”]", '', val)  # 引号
        return val.strip()

    # 标准化所有 key_set 中的值
    standardized_keys = set(tuple(standardize(x) for x in key) for key in key_set)

    # st.write(f"🟡 标黄匹配日志 - Sheet: {ws.title}")

    for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
        key_raw = [ws.cell(row=row, column=col).value for col in key_cols]
        key = tuple(standardize(v) for v in key_raw)
        display_key = tuple(key_raw)  # 用原始值用于日志输出
        # st.write(f"第 {row} 行匹配尝试: {display_key}")
        if key in standardized_keys:
            # st.write(f"✅ 第 {row} 行匹配成功: {display_key}")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill
        # else:
            # st.write(f"❌ 第 {row} 行未匹配: {display_key}")
