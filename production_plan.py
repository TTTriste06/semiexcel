from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font


HEADER_TEMPLATE = [
    "销售数量", "销售金额", "成品投单计划", "半成品投单计划", "投单计划周期", 
    "成品可行投单", "半成品可行投单", "成品实际投单", "半成品实际投单", 
    "回货计划", "回货计划调整", "PC回货计划", "回货实际"
]

def insert_repeated_headers(ws, start_col: int, start_month: int, end_month: int):
    """
    插入合并的月份 header 与重复的 field header。
    """
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    bold_center = Alignment(horizontal="center", vertical="center")
    font = Font(bold=True)

    current_col = start_col
    for m in range(start_month, end_month + 1):
        # 合并单元格（上层月份名）
        merge_range = f"{get_column_letter(current_col)}1:{get_column_letter(current_col + len(HEADER_TEMPLATE) - 1)}1"
        ws.merge_cells(merge_range)
        ws.cell(row=1, column=current_col).value = f"{m}月"
        ws.cell(row=1, column=current_col).fill = yellow_fill
        ws.cell(row=1, column=current_col).alignment = bold_center
        ws.cell(row=1, column=current_col).font = font

        # 第二行写入模板字段
        for i, header in enumerate(HEADER_TEMPLATE):
            cell = ws.cell(row=2, column=current_col + i)
            cell.value = header
            cell.fill = yellow_fill
            cell.alignment = bold_center
            cell.font = font

        current_col += len(HEADER_TEMPLATE)













def add_colored_monthly_plan_headers(ws, start_col: int, start_date: datetime, pivot_unfulfilled) -> int:
    """
    向“产品生产计划” Sheet 添加多月字段组表头（合并单元格，字段名，彩色背景）。
    
    参数：
    - ws: openpyxl Worksheet 对象
    - start_col: 从第几列开始插入月份组（通常是已有字段后的第1列）
    - start_date: 用户在界面中选择的排产起始月份（datetime 对象）
    - pivot_unfulfilled: 未交订单透视表，用于提取最大月份字段
    
    返回：
    - 最终写入结束的列号（用于后续插入数据）
    """

    # ✅ 每月字段列名
    monthly_fields = [
        "成品投单计划", "投单计划调整", "半成品投单计划",
        "成品可行投单", "半成品可行投单", "成品实际投单",
        "回货计划", "回货实际"
    ]

    # ✅ 每月对应背景色（12个以内自动轮换）
    month_colors = [
        "FFFACD", "FFDAB9", "FFE4E1", "87CEFA", "D8BFD8", "FFC0CB"
    ]

    # ✅ 提取最大月份（从未交订单列名中解析）
    month_pattern = re.compile(r"(\d{4})年(\d{1,2})月.*未交订单数量")
    max_month = None
    for col in pivot_unfulfilled.columns:
        match = month_pattern.match(col)
        if match:
            year, month = int(match.group(1)), int(match.group(2))
            dt = datetime(year, month, 1)
            if not max_month or dt > max_month:
                max_month = dt
    end_date = max_month if max_month else start_date + relativedelta(months=6)

    # ✅ 开始写入月份组表头
    current_col = start_col - 1
    month_index = 0
    while start_date <= end_date:
        fill_color = PatternFill("solid", fgColor=month_colors[month_index % len(month_colors)])
        month_name = start_date.strftime("%-m月")

        # 第一行：合并单元格并写月份名
        ws.merge_cells(
            start_row=1, start_column=current_col,
            end_row=1, end_column=current_col + len(monthly_fields) - 1
        )
        cell = ws.cell(row=1, column=current_col)
        cell.value = month_name
        cell.fill = fill_color
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 第二行：写字段名
        for offset, field in enumerate(monthly_fields):
            col = current_col + offset
            ws.cell(row=2, column=col, value=field)
            ws.cell(row=2, column=col).fill = fill_color

        # 下一月
        current_col += len(monthly_fields)
        start_date += relativedelta(months=1)
        month_index += 1

    # ✅ 自动调整新列区域的列宽
    adjust_column_width_ws(ws)


     # ✅ 设置黑边框


    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for row in [1, 2]:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = black_border

    return current_col  # 返回最后写入的列号



def calculate_first_month_plan(df_plan: pd.DataFrame, summary_df: pd.DataFrame, first_month: datetime) -> pd.DataFrame:
    """
    计算第一个月的“成品投单计划”列，考虑安全库存 + max(预测, 订单) + ... - 库存 - 在制
    """


    # 🔁 构造字段名
    month1_str = first_month.strftime("%Y年%m月")
    month2_str = (first_month + relativedelta(months=1)).strftime("%Y年%m月")

    # ✅ 计算用于列名的字符串
    col_forecast_1 = "6月预测"
    col_order_1 = "未交订单数量_2025-06"
    col_forecast_2 = "7月预测"
    col_order_2 = "未交订单数量_2025-07"

    col_inv = " InvPart"
    col_finished_1 = "数量_成品仓"
    col_in_progress = "成品在制"

    

    # ✅ 保证字段存在
    needed_columns = [col_forecast_1, col_order_1, col_forecast_2, col_order_2,
                      col_inv, col_finished_1, col_in_progress]
    for col in needed_columns:
        if col not in summary_df.columns:
            summary_df[col] = 0

    # ✅ 强制类型转为 float 并填 0
    def safe_float(series):
        return pd.to_numeric(series, errors="coerce").fillna(0)

    part_inv = safe_float(summary_df[col_inv])
    forecast_1 = safe_float(summary_df[col_forecast_1])
    order_1 = safe_float(summary_df[col_order_1])
    forecast_2 = safe_float(summary_df[col_forecast_2])
    order_2 = safe_float(summary_df[col_order_2])
    finished_inventory = safe_float(summary_df[col_finished_1])
    in_progress = safe_float(summary_df[col_in_progress])

    st.write(part_inv,forecast_1,order_1,forecast_2,order_2,finished_inventory,in_progress)

    # ✅ 按照公式计算
    plan = part_inv + pd.DataFrame({"a": forecast_1, "b": order_1}).max(axis=1) + \
           pd.DataFrame({"a": forecast_2, "b": order_2}).max(axis=1) - \
           finished_inventory - in_progress
    st.write(plan)

    # ✅ clip 保底 + 转 int
    plan = plan.clip(lower=0).round().astype(int)

    # ✅ 强制写入 df_plan 的第 8 列（H列）
    if df_plan.shape[1] >= 8:
        df_plan.iloc[:, 7] = plan  # 第 8 列是 index 7，对应 H 列
    else:
        raise ValueError("❌ df_plan 的列数不足 8 列，无法写入 H 列（成品投单计划）")
    return df_plan



def highlight_plan_column(ws, safe_col_name="安全库存", plan_col_name="成品投单计划", header_row=2):
    """
    给“成品投单计划”列应用条件着色：
    - 红色：值 < 0
    - 黄色：0 ≤ 值 < 安全库存
    - 橙色：值 > 2 × 安全库存

    参数：
    - ws: openpyxl Worksheet 对象
    - safe_col_name: 安全库存列名
    - plan_col_name: 成品投单计划列名
    - header_row: 表头所在行（默认第 2 行）
    """
    # ✅ 查找列索引
    def find_col_index(col_name):
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=col).value).strip()
            if val == col_name:
                return col
        return None

    col_safe = find_col_index(safe_col_name)
    col_plan = find_col_index(plan_col_name)

    if col_safe is None or col_plan is None:
        raise ValueError(f"❌ 找不到 '{safe_col_name}' 或 '{plan_col_name}' 列")

    # ✅ 设置填充颜色
    red_fill = PatternFill(fill_type="solid", fgColor="FF0000")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    orange_fill = PatternFill(fill_type="solid", fgColor="FFA500")

    # ✅ 从 header_row + 1 开始逐行处理
    for row in range(header_row + 1, ws.max_row + 1):
        try:
            val_plan = ws.cell(row=row, column=col_plan).value
            val_safe = ws.cell(row=row, column=col_safe).value
            val_plan = float(val_plan) if val_plan not in [None, ""] else 0
            val_safe = float(val_safe) if val_safe not in [None, ""] else 0

            cell = ws.cell(row=row, column=col_plan)

            if val_plan < 0:
                cell.fill = red_fill
            elif val_plan < val_safe:
                cell.fill = yellow_fill
            elif val_plan > 2 * val_safe:
                cell.fill = orange_fill

        except Exception as e:
            # 跳过不能转换的行
            continue

