import os
import io
import re
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta
from config import CONFIG
from excel_utils import (
    adjust_column_width,
    clean_df,
    merge_header_for_summary, 
    mark_unmatched_keys_on_sheet,
    mark_keys_on_sheet,
    merge_duplicate_product_names,
    merge_duplicate_rows_by_key,
    clean_key_fields,
    mark_unmatched_keys_on_name,
    reorder_summary_columns,
    clear_nan_cells,
    get_column_index_by_name
)
from mapping_utils import (
    apply_mapping_and_merge,
    apply_extended_substitute_mapping
)
from month_selector import process_history_columns
from summary import (
    merge_safety_inventory,
    append_unfulfilled_summary_columns,
    append_forecast_to_summary,
    merge_finished_inventory,
    append_product_in_progress
)
from append_summary import append_forecast_unmatched_to_summary_by_keys
from production_plan import insert_repeated_headers


FIELD_MAPPINGS = {
    "赛卓-未交订单": {"规格": "规格", "品名": "品名", "晶圆品名": "晶圆品名"},
    "赛卓-成品在制": {"规格": "产品规格", "品名": "产品品名", "晶圆品名": "晶圆型号"},
    "赛卓-成品库存": {"规格": "规格", "品名": "品名", "晶圆品名": "WAFER品名"},
    "赛卓-安全库存": {"规格": "OrderInformation", "品名": "ProductionNO.", "晶圆品名": "WaferID"},
    "赛卓-预测": {"品名": "生产料号"},
    "赛卓-到货明细.xlsx": {"品名": "品名"},
    "赛卓-下单明细.xlsx": {"品名": "回货明细_回货品名"},
    "赛卓-销货明细.xlsx": {"品名": "品名"}
}


class PivotProcessor:
    def process(self, uploaded_files: dict, output_buffer, additional_sheets: dict = None):
        df_finished = pd.DataFrame()
        product_in_progress = pd.DataFrame()
        df_unfulfilled = pd.DataFrame()

        unmatched_safety = []
        unmatched_unfulfilled = []
        unmatched_forecast = []
        unmatched_finished = []
        unmatched_in_progress = []

        key_unfulfilled = []
        key_finished = []
        key_in_progress = []

        mapping_df = additional_sheets.get("赛卓-新旧料号", pd.DataFrame())
        
        
        all_mapped_keys = set()

        # 清洗 additional_sheets 中的所有 nan 字符串
        for name in ["赛卓-预测", "赛卓-安全库存", "赛卓-新旧料号"]:
            if name in additional_sheets:
                df = additional_sheets[name]
                df = df.fillna("")  # 替换真正的 NaN
                df = df.applymap(lambda x: "" if str(x).strip().lower() == "nan" else str(x).strip() if isinstance(x, str) else x)
                additional_sheets[name] = df  # 更新为清洗后的 df

        # 在 PivotProcessor.process 内部，写 Excel 之前：
        # 检查是否有表含有字符串 "nan"
        for name, df in additional_sheets.items():
            if (df.astype(str).applymap(lambda x: x.lower() == "nan")).any().any():
                st.warning(f"⚠️ 表 `{name}` 中含有字符串 'nan'，请确认是否清洗干净")



        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            for filename, file_obj in uploaded_files.items():
                try:
                    df = pd.read_excel(file_obj)
                    df = clean_df(df)
                    config = CONFIG["pivot_config"].get(filename)
                    if not config:
                        st.warning(f"⚠️ 跳过未配置的文件：{filename}")
                        continue

                    sheet_name = filename.replace(".xlsx", "")

                    if sheet_name in FIELD_MAPPINGS and not mapping_df.empty:
                        mapping_df.columns = [
                            "旧规格", "旧品名", "旧晶圆品名",
                            "新规格", "新品名", "新晶圆品名",
                            "封装厂", "PC", "半成品", "备注",
                            "替代规格1", "替代品名1", "替代晶圆1", 
                            "替代规格2", "替代品名2", "替代晶圆2", 
                            "替代规格3", "替代品名3", "替代晶圆3",
                            "替代规格4", "替代品名4", "替代晶圆4"
                        ] + list(mapping_df.columns[22:])
                        st.success(f"✅ `{sheet_name}` 正在进行新旧料号替换...")

                        df, mapped_keys = apply_mapping_and_merge(df, mapping_df, FIELD_MAPPINGS[sheet_name])
                        df, keys_sub = apply_extended_substitute_mapping(df, mapping_df, FIELD_MAPPINGS[sheet_name], None)
                        df = clean_key_fields(df, FIELD_MAPPINGS[sheet_name])
                        df = merge_duplicate_rows_by_key(df, FIELD_MAPPINGS[sheet_name])
                        all_mapped_keys.update(mapped_keys)

                        if sheet_name == "赛卓-未交订单":
                            key_unfulfilled = mapped_keys
                        elif sheet_name == "赛卓-成品库存":
                            key_finished = mapped_keys
                        elif sheet_name == "赛卓-成品在制":
                            key_in_progress = mapped_keys

                    if "date_format" in config:
                        df = self._process_date_column(df, config["columns"], config["date_format"])

                    pivoted = self._create_pivot(df, config)
                    pivoted.to_excel(writer, sheet_name=sheet_name, index=False)
                    adjust_column_width(writer, sheet_name, pivoted)

                    if sheet_name == "赛卓-未交订单":
                        df_unfulfilled = df
                        pivot_unfulfilled = pivoted
                    elif sheet_name == "赛卓-成品库存":
                        df_finished = pivoted
                    elif sheet_name == "赛卓-成品在制":
                        product_in_progress = pivoted

                except Exception as e:
                    st.error(f"❌ 文件 `{filename}` 处理失败: {e}")

            if df_unfulfilled.empty:
                st.error("❌ 缺少未交订单数据，无法构建汇总")
                return

            summary_preview = df_unfulfilled[["晶圆品名", "规格", "品名"]].drop_duplicates().reset_index(drop=True)
                
            try:
                if "赛卓-预测" in additional_sheets:
                    forecast_df = additional_sheets["赛卓-预测"]
                    forecast_df = clean_df(forecast_df)
                    forecast_df, keys_main = apply_mapping_and_merge(forecast_df, mapping_df, FIELD_MAPPINGS["赛卓-预测"])
                    ## forecast_df, keys_sub = apply_extended_substitute_mapping(forecast_df, mapping_df, FIELD_MAPPINGS["赛卓-预测"], keys_main)
                    # forecast_df = merge_duplicate_rows_by_key(forecast_df, FIELD_MAPPINGS["赛卓-预测"])
                    # all_mapped_keys.update(keys_main)
                    # all_mapped_keys.update(keys_sub)
                    summary_preview, unmatched_forecast = append_forecast_to_summary(summary_preview, forecast_df)
                    st.success("✅ 已合并预测数据")
                    
                    
                    # 添加未匹配的预测项
                    summary_preview = append_forecast_unmatched_to_summary_by_keys(summary_preview, forecast_df)
                    st.success("✅ 已添加未匹配的预测项至汇总表")

                
                if "赛卓-安全库存" in additional_sheets:
                    df_safety = additional_sheets["赛卓-安全库存"]
                    df_safety = clean_df(df_safety)
                    df_safety, keys_main = apply_mapping_and_merge(df_safety, mapping_df, FIELD_MAPPINGS["赛卓-安全库存"])
                    df_safety, keys_sub = apply_extended_substitute_mapping(df_safety, mapping_df, FIELD_MAPPINGS["赛卓-安全库存"], keys_main)
                    df_safety = merge_duplicate_rows_by_key(df_safety, FIELD_MAPPINGS["赛卓-安全库存"])
                    # all_mapped_keys.update(keys_main)
                    # all_mapped_keys.update(keys_sub)
                    summary_preview, unmatched_safety = merge_safety_inventory(summary_preview, df_safety)
                    st.success("✅ 已合并安全库存")
                    
                summary_preview, unmatched_unfulfilled = append_unfulfilled_summary_columns(summary_preview, pivot_unfulfilled)
                st.success("✅ 已合并未交订单")
                
                # ✅ 提取最大月份字段
                month_pattern = re.compile(r"(\d{4})年(\d{1,2})月.*未交订单数量")
                max_month = None
                
                for col in pivot_unfulfilled.columns:
                    match = month_pattern.match(col)
                    if match:
                        year, month = int(match.group(1)), int(match.group(2))
                        dt = datetime(year, month, 1)
                        if not max_month or dt > max_month:
                            max_month = dt
                
                if max_month:
                    end_date = max_month
                else:
                    end_date = datetime.today() + relativedelta(months=6)  # 默认未来 6 个月
                

                if not df_finished.empty:
                    summary_preview, unmatched_finished = merge_finished_inventory(summary_preview, df_finished)
                    st.success("✅ 已合并成品库存")

                if not product_in_progress.empty:
                    summary_preview, unmatched_in_progress = append_product_in_progress(summary_preview, product_in_progress, mapping_df)
                    st.success("✅ 已合并成品在制")


                summary_preview = clean_df(summary_preview)
                summary_preview = summary_preview.drop_duplicates(subset=["晶圆品名", "规格", "品名"]).reset_index(drop=True)
                summary_preview = merge_duplicate_product_names(summary_preview)
                summary_preview = reorder_summary_columns(summary_preview)


                HEADER_TEMPLATE = [
                    "销售数量", "销售金额", "成品投单计划", "半成品投单计划", "投单计划调整",
                    "成品可行投单", "半成品可行投单", "成品实际投单", "半成品实际投单",
                    "回货计划", "回货计划调整", "PC回货计划", "回货实际"
                ]


                # 在保存 summary_preview 前插入：
                today_month = datetime.today().month
                month_pattern = re.compile(r"(\d{1,2})月预测")
                forecast_months = []
                
                for col in summary_preview.columns:
                    match = month_pattern.match(str(col))
                    if match:
                        forecast_months.append(int(match.group(1)))

                st.write(forecast_months)
                
                # 确定添加月份范围
                start_month = today_month
                end_month = max(forecast_months) - 1 if forecast_months else start_month

                # ✅ 在 summary_preview 中添加每月字段列（全部初始化为空或0）
                for m in range(start_month, end_month + 1):
                    for header in HEADER_TEMPLATE:
                        new_col = f"{m}_{header}"
                        summary_preview[new_col] = ""



                def safe_col(df, col):
                    # 返回确保是 float 的 Series，字符串将被转为 NaN，再用 0 替代
                    return pd.to_numeric(df[col], errors="coerce").fillna(0) if col in df.columns else pd.Series(0, index=df.index)




                df_plan = pd.DataFrame(index=summary_preview.index)

                for idx, month in enumerate(forecast_months[:-1]):  # 最后一个月不生成
                    this_month = f"{month}月"
                    next_month = f"{forecast_months[idx + 1]}月"
                    prev_month = f"{forecast_months[idx - 1]}月" if idx > 0 else None
                
                    # 构造字段名
                    col_forecast_this = f"{month}月预测"
                    col_order_this = f"未交订单数量_2025-{month}"
                    col_forecast_next = f"{forecast_months[idx + 1]}月预测"
                    col_order_next = f"未交订单数量_2025-{forecast_months[idx + 1]}"
                    col_target = f"{this_month}_成品投单计划"
                    col_actual_prod = f"{this_month}_成品实际投单"
                    col_target_prev = f"{prev_month}_成品投单计划" if prev_month else None
                
                    if idx == 0:
                        # 第一个月：特殊算法
                        df_plan[col_target] = (
                            safe_col(summary_preview, "InvPart") +
                            pd.DataFrame({
                                "f": safe_col(summary_preview, col_forecast_this),
                                "o": safe_col(summary_preview, col_order_this)
                            }).max(axis=1) +
                            pd.DataFrame({
                                "f": safe_col(summary_preview, col_forecast_next),
                                "o": safe_col(summary_preview, col_order_next)
                            }).max(axis=1) -
                            safe_col(summary_preview, "数量_成品仓") -
                            safe_col(summary_preview, "成品在制")
                        )
                    else:
                        df_plan[col_target] = (
                            pd.DataFrame({
                                "f": safe_col(summary_preview, col_forecast_next),
                                "o": safe_col(summary_preview, col_order_next)
                            }).max(axis=1) +
                            (safe_col(df_plan, col_target_prev) - safe_col(summary_preview, col_actual_prod))
                        )



    
                
                # ✅ 只选 summary 中的“成品投单计划”列（排除半成品）
                plan_cols_in_summary = [col for col in summary_preview.columns if "成品投单计划" in col and "半成品" not in col]
                
                # ✅ 数量校验
                if len(plan_cols_in_summary) != df_plan.shape[1]:
                    st.error(f"❌ 写入失败：df_plan 有 {df_plan.shape[1]} 列，summary 中有 {len(plan_cols_in_summary)} 个 '成品投单计划' 列")
                else:
                    # ✅ 将 df_plan 的列按顺序填入 summary_preview
                    for i, col in enumerate(plan_cols_in_summary):
                        summary_preview[col] = df_plan.iloc[:, i]
                
                    st.success("✅ 成品投单计划已写入 summary_preview")






                

                df_semi_plan = pd.DataFrame(index=summary_preview.index)

                plan_cols = df_plan.columns.tolist()  # 如 ['5月_成品投单计划', '6月_成品投单计划', ...]
                
                for i, col in enumerate(plan_cols):
                    if i == 0:
                        # 第一个月：成品投单计划 - 半成品在制
                        df_semi_plan[col.replace("成品投单计划", "半成品投单计划")] = (
                            df_plan[col] - safe_col(summary_preview, "半成品在制")
                        )
                    else:
                        df_semi_plan[col.replace("成品投单计划", "半成品投单计划")] = (
                            0
                        )




                # 回货实际
                # ✅ 提取原始数据
                df_arrival = additional_sheets.get("赛卓-到货明细", pd.DataFrame())
                df_arrival = df_arrival[["到货日期", "品名", "允收数量"]].copy()
                
                # ✅ （可选映射）跳过暂不启用
                # df_arrival, keys_main = apply_mapping_and_merge(df_arrival, mapping_df, FIELD_MAPPINGS["赛卓-到货明细"])
                # df_arrival, _ = apply_extended_substitute_mapping(df_arrival, mapping_df, FIELD_MAPPINGS["赛卓-到货明细"], keys_main)
                
    
                
                # ✅ 清理：只保留汇总中存在的品名
                valid_names = set(summary_preview["品名"].astype(str))
                df_arrival["品名"] = df_arrival["品名"].astype(str)
                df_arrival = df_arrival[df_arrival["品名"].isin(valid_names)]
                
                # ✅ 初始化结果表，第一列来自 summary_preview 的品名（跳过第一行 header）
                arrival_by_month = pd.DataFrame()
                arrival_by_month["品名"] = summary_preview.loc[1:, "品名"].astype(str).reset_index(drop=True)
                
                for m in forecast_months:
                    col_name = f"{m}月到货数量"
                    arrival_by_month[col_name] = 0  # 初始化为 0
                
                # ✅ 计算每条记录的到货月份
                df_arrival["到货月份"] = pd.to_datetime(df_arrival["到货日期"], errors="coerce").dt.month
                
                # ✅ 遍历并累加到目标月份列中
                for idx, row in df_arrival.iterrows():
                    part = row["品名"]
                    qty = row["允收数量"]
                    month = row["到货月份"]
                    if month in forecast_months:
                        col = f"{month}月到货数量"
                        match_idx = arrival_by_month[arrival_by_month["品名"] == part].index
                        if not match_idx.empty:
                            arrival_by_month.loc[match_idx[0], col] += qty


                back_cols_in_summary = [col for col in summary_preview.columns if "回货实际" in col]
                

                # ✅ 按顺序填入 summary_preview
                for i, col in enumerate(back_cols_in_summary):
                    summary_preview[col] = arrival_by_month.iloc[:, i+1]
            
                st.success("✅ 回货实际已写入 summary_preview")



                # 销货数量和销货金额
                # ✅ 提取销货明细原始数据
                df_sales = additional_sheets.get("赛卓-销货明细", pd.DataFrame())
                df_sales = df_sales[["交易日期", "品名", "数量", "原币金额"]].copy()
                
                # ✅ 清理：只保留出现在 summary_preview 中的品名
                valid_names = set(summary_preview["品名"].astype(str))
                df_sales["品名"] = df_sales["品名"].astype(str)
                df_sales = df_sales[df_sales["品名"].isin(valid_names)]
                
                # ✅ 初始化两个结果表（以 summary_preview 的品名为基准，跳过 header 行）
                sales_qty_by_month = pd.DataFrame()
                sales_amt_by_month = pd.DataFrame()
                sales_qty_by_month["品名"] = summary_preview.loc[1:, "品名"].astype(str).reset_index(drop=True)
                sales_amt_by_month["品名"] = summary_preview.loc[1:, "品名"].astype(str).reset_index(drop=True)
                
                for m in forecast_months:
                    col_qty = f"{m}月销售数量"
                    col_amt = f"{m}月销售金额"
                    sales_qty_by_month[col_qty] = 0
                    sales_amt_by_month[col_amt] = 0
                
                # ✅ 提取月份
                df_sales["销售月份"] = pd.to_datetime(df_sales["交易日期"], errors="coerce").dt.month
                
                # ✅ 累加销售数据
                for idx, row in df_sales.iterrows():
                    part = row["品名"]
                    qty = row["数量"]
                    amt = row["原币金额"]
                    month = row["销售月份"]
                    if month in forecast_months:
                        col_qty = f"{month}月销售数量"
                        col_amt = f"{month}月销售金额"
                        match_idx = sales_qty_by_month[sales_qty_by_month["品名"] == part].index
                        if not match_idx.empty:
                            sales_qty_by_month.loc[match_idx[0], col_qty] += qty
                            sales_amt_by_month.loc[match_idx[0], col_amt] += amt
                
                # ✅ 写入汇总表 summary_preview（跳过 header）
                sales_qty_cols_in_summary = [col for col in summary_preview.columns if "销售数量" in col]
                sales_amt_cols_in_summary = [col for col in summary_preview.columns if "销售金额" in col]
                
                for i, col in enumerate(sales_qty_cols_in_summary):
                    if i + 1 < sales_qty_by_month.shape[1]:
                        summary_preview.loc[1:, col] = sales_qty_by_month.iloc[:, i + 1].values
                
                for i, col in enumerate(sales_amt_cols_in_summary):
                    if i + 1 < sales_amt_by_month.shape[1]:
                        summary_preview.loc[1:, col] = sales_amt_by_month.iloc[:, i + 1].values
                
                st.success("✅ 销售数量与销售金额已写入 summary_preview")






                # 成品实际投单
                # ✅ 提取下单明细原始数据
                df_order = additional_sheets.get("赛卓-下单明细", pd.DataFrame())
                df_order = df_order[["下单日期", "回货明细_回货品名", "回货明细_回货数量"]].copy()
                
                # ✅ 清理：只保留 summary 中有的品名
                valid_names = set(summary_preview["品名"].astype(str))
                df_order["回货明细_回货品名"] = df_order["回货明细_回货品名"].astype(str)
                df_order = df_order[df_order["回货明细_回货品名"].isin(valid_names)]
                
                # ✅ 初始化结果表
                order_plan_by_month = pd.DataFrame()
                order_plan_by_month["品名"] = summary_preview.loc[1:, "品名"].astype(str).reset_index(drop=True)
                
                for m in forecast_months:
                    col_name = f"{m}月成品实际投单"
                    order_plan_by_month[col_name] = 0
                
                # ✅ 提取月份
                df_order["下单月份"] = pd.to_datetime(df_order["下单日期"], errors="coerce").dt.month
                
                # ✅ 累加每条记录的数量到对应月份
                for idx, row in df_order.iterrows():
                    part = row["回货明细_回货品名"]
                    qty = row["回货明细_回货数量"]
                    month = row["下单月份"]
                    if month in forecast_months:
                        col = f"{month}月成品实际投单"
                        match_idx = order_plan_by_month[order_plan_by_month["品名"] == part].index
                        if not match_idx.empty:
                            order_plan_by_month.loc[match_idx[0], col] += qty
                
                # ✅ 找出 summary 中对应的列并填入
                order_cols_in_summary = [col for col in summary_preview.columns if "成品实际投单" in col and "半成品" not in col]
                
                for i, col in enumerate(order_cols_in_summary):
                    if i + 1 < order_plan_by_month.shape[1]:
                        summary_preview.loc[1:, col] = order_plan_by_month.iloc[:, i + 1].values
                
                st.success("✅ 成品实际投单已写入 summary_preview")



                
               
     



            except Exception as e:
                st.error(f"❌ 汇总数据合并失败: {e}")
                return

            summary_preview.to_excel(writer, sheet_name="汇总", index=False)
            adjust_column_width(writer, "汇总", summary_preview)



            ws = writer.sheets["汇总"]

            
            # 半成品投单计划
            semi_plan_cols_in_summary = [col for col in summary_preview.columns if "半成品投单计划" in col]

            
            for i, col in enumerate(semi_plan_cols_in_summary):
                col_idx = summary_preview.columns.get_loc(col) + 1  # 1-based Excel column index
                col_letter = get_column_letter(col_idx)
                
            
                for row in range(3, len(summary_preview) + 3):  # 数据从第3行开始
                    cell = ws.cell(row=row, column=col_idx)
            
                    if i == 0:
                        # 第一个月：填入真实数值
                        cell.value = df_semi_plan.iloc[row - 3, 0]
                    else:
                        # 后续月份：填入公式
                        prev_col_letter = get_column_letter(col_idx - 1)
                        col_13_back = get_column_letter(col_idx - 13)
                        col_8_back = get_column_letter(col_idx - 8)
                        formula = f"={prev_col_letter}{row} + ({col_13_back}{row} - {col_8_back}{row})"
                        
                        cell.value = formula


            # 投单计划调整
            adjust_plan_cols_in_summary = [col for col in summary_preview.columns if "投单计划调整" in col]
            
            for i, col in enumerate(adjust_plan_cols_in_summary):
                col_idx = summary_preview.columns.get_loc(col) + 1  # 1-based Excel column index
                col_letter = get_column_letter(col_idx)
                
            
                for row in range(3, len(summary_preview) + 3):  # 数据从第3行开始
                    cell = ws.cell(row=row, column=col_idx)
            
                    if i == 0:
                        cell.value = ""
                    else:
                        # 后续月份：填入公式
                        prev_col_letter = get_column_letter(col_idx - 2)
                        col_13_back = get_column_letter(col_idx - 15)
                        col_8_back = get_column_letter(col_idx - 12)
                        formula = f"={prev_col_letter}{row} + ({col_13_back}{row} - {col_8_back}{row})"
                        cell.value = formula



            # 回货计划
            return_plan_cols_in_summary = [col for col in summary_preview.columns if "回货计划" in col and "调整" not in col]
            
            for i, col in enumerate(return_plan_cols_in_summary):
                col_idx = summary_preview.columns.get_loc(col) + 1  # 1-based Excel column index
                col_letter = get_column_letter(col_idx)
                
            
                for row in range(3, len(summary_preview) + 3):  # 数据从第3行开始
                    cell = ws.cell(row=row, column=col_idx)
            
                    if i == 0:
                        cell.value = ""
                    else:
                        # 后续月份：填入公式
                        prev_col_letter = get_column_letter(col_idx - 18)
                        formula = f"={prev_col_letter}{row}"
                        cell.value = formula


            # 回货计划调整
            adjust_return_plan_cols_in_summary = [col for col in summary_preview.columns if "回货计划调整" in col]
            
            for i, col in enumerate(adjust_return_plan_cols_in_summary):
                col_idx = summary_preview.columns.get_loc(col) + 1  # 1-based Excel column index
                col_letter = get_column_letter(col_idx)
                
            
                for row in range(3, len(summary_preview) + 3):  # 数据从第3行开始
                    cell = ws.cell(row=row, column=col_idx)
            
                    if i == 0:
                        cell.value = ""
                    else:
                        # 后续月份：填入公式
                        prev_col_letter = get_column_letter(col_idx - 1)
                        col_13_back = get_column_letter(col_idx - 16)
                        col_8_back = get_column_letter(col_idx - 19)
                        formula = f"={prev_col_letter}{row} + ({col_13_back}{row} - {col_8_back}{row})"
                        
                        cell.value = formula
            
            



            header_row = list(summary_preview.columns)
            unfulfilled_cols = [col for col in header_row if "未交订单数量" in col or col in ("总未交订单", "历史未交订单数量")]
            forecast_cols = [col for col in header_row if "预测" in col]
            finished_cols = [col for col in header_row if col in ("数量_HOLD仓", "数量_成品仓", "数量_半成品仓")]



            

            merge_header_for_summary(
                ws, summary_preview,
                {
                    "安全库存": (" InvWaf", " InvPart"),
                    "未交订单": (unfulfilled_cols[0], unfulfilled_cols[-1]),
                    "预测": (forecast_cols[0], forecast_cols[-1]) if forecast_cols else ("", ""),
                    "成品库存": (finished_cols[0], finished_cols[-1]) if finished_cols else ("", ""),
                    "成品在制": ("成品在制", "半成品在制")
                }
            )

            for key, df in additional_sheets.items():
                df.to_excel(writer, sheet_name=key, index=False)
                adjust_column_width(writer, key, df)

            # 每个 sheet 中用于标记的字段名（目标列）及表头所在行（从 1 开始）
            sheet_field_config = {
                "赛卓-安全库存": {"field_name": "ProductionNO.", "header_row": 1},
                "赛卓-未交订单": {"field_name": "品名", "header_row": 1},
                "赛卓-预测": {"field_name": "生产料号", "header_row": 1},
                "汇总": {"field_name": "品名", "header_row": 2},  # 汇总表通常从第2行起才是字段行
                "赛卓-成品库存": {"field_name": "品名", "header_row": 1},
                "赛卓-成品在制": {"field_name": "产品品名", "header_row": 1},
            }

            sheet_key_mapping = {
                    "赛卓-安全库存": unmatched_safety,
                    "赛卓-未交订单": unmatched_unfulfilled,
                    "赛卓-预测": unmatched_forecast,
                    "赛卓-成品库存": unmatched_finished,
                    "赛卓-成品在制": unmatched_in_progress,
                }
                

            try:
                # 标红未匹配行
                for sheet_name, unmatched_keys in sheet_key_mapping.items():
                    if sheet_name in writer.sheets and sheet_name in sheet_field_config:
                        ws = writer.sheets[sheet_name]
                        config = sheet_field_config[sheet_name]
                        field_name = config["field_name"]
                        header_row = config["header_row"]
                        col_idx = get_column_index_by_name(ws, field_name, header_row)
                
                        if col_idx:
                            mark_unmatched_keys_on_name(ws, unmatched_keys, name_col=col_idx)
                        else:
                            st.warning(f"⚠️ `{sheet_name}` 中未找到字段 `{field_name}`，跳过未匹配标记")

                mark_unmatched_keys_on_name(writer.sheets["汇总"], unmatched_forecast, name_col=3)


                """
                标黄
                mark_keys_on_sheet(writer.sheets["汇总"], all_mapped_keys, (2, 3, 1))
                mark_keys_on_sheet(writer.sheets["赛卓-安全库存"], all_mapped_keys, (3, 5, 1))
                mark_keys_on_sheet(writer.sheets["赛卓-未交订单"], all_mapped_keys, (2, 3, 1))
                mark_keys_on_sheet(writer.sheets["赛卓-预测"], all_mapped_keys, (1, 2, 3))
                mark_keys_on_sheet(writer.sheets["赛卓-成品库存"], all_mapped_keys, (2, 3, 1))
                mark_keys_on_sheet(writer.sheets["赛卓-成品在制"], all_mapped_keys, (4, 5, 3))
                """

                st.success("✅ 已完成未匹配项标记")
            except Exception as e:
                st.warning(f"⚠️ 未匹配标记失败：{e}")



        # ✅ 替换这些指定表的 nan 值为空
        for sheet_name in ["赛卓-安全库存", "赛卓-预测", "赛卓-新旧料号", "汇总"]:
            if sheet_name in writer.sheets:
                clear_nan_cells(writer.sheets[sheet_name])


            output_buffer.seek(0)
        


    def _process_date_column(self, df, date_col, date_format):
        if pd.api.types.is_numeric_dtype(df[date_col]):
            df[date_col] = df[date_col].apply(self._excel_serial_to_date)
        else:
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

        new_col = f"{date_col}_年月"
        df[new_col] = df[date_col].dt.strftime(date_format)
        df[new_col] = df[new_col].fillna("未知日期")
        return df

    def _excel_serial_to_date(self, serial):
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(serial))
        except:
            return pd.NaT

    def _create_pivot(self, df, config):
        config = config.copy()
        if "date_format" in config:
            config["columns"] = f"{config['columns']}_年月"

        pivoted = pd.pivot_table(
            df,
            index=config["index"],
            columns=config["columns"],
            values=config["values"],
            aggfunc=config["aggfunc"],
            fill_value=0
        )

        pivoted.columns = [f"{col[0]}_{col[1]}" if isinstance(col, tuple) else str(col) for col in pivoted.columns]

        if pd.Series(pivoted.columns).duplicated().any():
            from pandas.io.parsers import ParserBase
            original_cols = pivoted.columns
            deduped_cols = ParserBase({'names': original_cols})._maybe_dedup_names(original_cols)
            pivoted.columns = deduped_cols

        pivoted = pivoted.reset_index()

        if CONFIG.get("selected_month") and config.get("values") and "未交订单数量" in config.get("values"):
            st.info(f"📅 合并历史数据至：{CONFIG['selected_month']}")
            pivoted = process_history_columns(pivoted, config, CONFIG["selected_month"])
        return pivoted
